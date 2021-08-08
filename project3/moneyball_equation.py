import openpyxl as excel
import readline

def main():

    # Welcome message:

    print("Welcome to the Moneyball Equation application!")

    while True:

        file_path = input("Please enter the absolute/relative path to your file" + \
                            " (enter 'q' to quit): ")

        if (file_path.lower() == "q"):

            print("Thank you for using Moneyball Equation application!\n\nQuitting...")
            exit(1)

        # Open excel sheet
        print("Openning file...")
        
        try:

            wb = excel.load_workbook(file_path)
            print("Opened file successfully!\n")
            break

        except FileNotFoundError:

            print("File not found! Please try again...\n(Note: Press \'q\' to quit)\n")
        
        except excel.utils.exceptions.InvalidFileException:

            print("File format not supported! Please try again...\n(Note: Press \'q\' to quit)\n")

    # Read Cell Data from excel sheet ws['CR'] where C == column and R == row

    # Reference:
    # Columns:
    # A ==> Team
    # B ==> Games Played
    # C ==> Runs Scored
    # D ==> Runs Allowed
    # E ==> Wins
    # F ==> Losses
    # G ==> W-L% (to be edited)
    # Rows:
    # 1 - the titles
    # 2-31 are valid teams

    # Iterate through each worksheet
    for name in wb.sheetnames:

        print(f"Writing to {name}...")

        # Enter the worksheet
        ws = wb[name]
        curr_percentage: float

        # For each row in the sheet
        for row_index in range(2, 32):

            # Find the current team's values
            wins = ws[f"E{row_index}"].value
            games_played = ws[f"B{row_index}"].value

            # Use the Pythagorean Expectation Formula
            percentage = (float(wins) / float(games_played))
            ws[f"G{row_index}"].value = percentage
            ws[f"G{row_index}"].number_format = "0.00%"

    # Save to a new file called similarily but insert '_edited'
    # before the .xlsx extension
    save_path: str = file_path[0:-5:1] + '_edited.xlsx'
    print(f"\nDone!\nSaving to {save_path}...")
    wb.save(save_path)

    # Close excel sheet for security reasons
    wb.close()

    print("\nExitting...\n")

    pass

# Driver of the file
if __name__ == "__main__":

    main()