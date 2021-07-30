import openpyxl as excel
import csv
import textwrap
import table
import blessed as b

# Constants
MAX_QUESTIONS: int = 100
IGNORED_HEADERS: list[str] = ["n correct", "n incorrect", "score"]

# Global variable
file_path: str

# Converts csv files into xlsx files
def csv_to_xlsx(path: str) -> excel.Workbook:

    wb = excel.Workbook()
    ws = wb.active

    with open(path) as file:
        reader = csv.reader(file, delimiter=',')
        for row in reader:
            ws.append(row)

    wb.save(path := path.replace(".csv", ".xlsx"))
    wb.close()

    return excel.load_workbook(path)

# Opens a file
def open_file() -> excel.Workbook:

    file_path: str

    while True:

        try:
            
            print("Please enter the absolute/relative path to the excel sheet (Press 'q' to quit): ", end='')
            file_path = input()

            if file_path.lower() == "q":

                return None

            wb =  excel.load_workbook(file_path)

            return wb

        except FileNotFoundError:

            print("File does not exit.\nPlease try again\n")

        except excel.utils.exceptions.InvalidFileException:

            return csv_to_xlsx(file_path)

# Finds the first question's column in the Excel Sheet
def find_start(ws) -> int:

    # Counter variables    
    col_count: int = 1
    curr_cell: str = ""

    # Look for where the question column starts
    while (curr_cell := ws["" + excel.utils.get_column_letter(col_count) + "1"].value):

        col_count += 1

        if (curr_cell == "attempt"):

            return col_count

    return None

# Finds the maximum number of students/answers to check for
def get_max_students(ws) -> int:

    # Counter variables
    num_of_students: int = 0
    row_num: int = 2

    # Find the max number of responses to expect
    while (ws["A" + str(row_num)].value):

        num_of_students += 1
        row_num += 1

    return num_of_students

# Parse questions from the excel sheet starting from the current column
def get_questions(ws, data_start, num_of_students) -> list[list[str]]:

    # Counter/Flag variables
    col_num: int = data_start
    row_num: int = 1
    break_now: bool = False
    curr_index: int = -1

    question_group: list[str] = []
    question_list: list[list[str]] = [question_group]

    while (col_num <= data_start + MAX_QUESTIONS * 2):

        if break_now:
            
            break

        for row_num in range(1, num_of_students + 1):

            curr_cell = ws[f"{excel.utils.get_column_letter(col_num)}{row_num}"].value

            if row_num == 1 and curr_cell.replace(".", "").isdecimal():

                break

            elif row_num == 1:

                question_group = []
                question_list.append(question_group)
                curr_index += 1

            if not curr_cell:

                continue

            elif curr_cell in IGNORED_HEADERS:
                
                break_now = True
                break

            curr_cell = curr_cell.replace("\n", "").encode("ascii", "ignore").decode()
            question_group.append(curr_cell)

        col_num += 1

    return question_list

def print_rows(term: b.Terminal, rows: list[str], curr: int) -> None:

    term.clear()

    bound = (curr + term.height - 2) # if (curr + term.height - 2 < len(rows)) else len(rows)

    for index in range(curr, bound):

        print(rows[index])

    pass
    
# The driver of the program
def main() -> None:

    # Welcome msg
    print("Welcome to my Canvas Feedback Reader!")

    # Prepare data
    canvas_wb = open_file()
    
    # Quit if no canvas csv file was provided (a.k.a selected to quit)
    if not canvas_wb:

        print("Thank you for using Canvas Feedback Reader!")
        exit(0)

    # Select the current worksheet
    canvas_ws = canvas_wb.active
    
    # Error checking for in case
    if not canvas_ws:

        print("ERROR: EMPTY WORKSHEET, ABORTING PROGRAM")
        exit(1)

    # Find the start of the questions in the excel sheet
    data_start = find_start(canvas_ws)

    # Error checking if there was no question column found
    if data_start < 0:

        print("ERROR: NO DATA TO READ, ABORTING PROGRAM")
        exit(1)

    num_of_students = get_max_students(canvas_ws)

    if (num_of_students < 1):

        print("ERROR: No Student Responses found!")
        exit(1)

    # Lists to store the questions collected
    # in the format:
    #
    # [[question1, answer1, answer2, ...], [question2, answer1, answer2], [...]]
    question_list = get_questions(canvas_ws, data_start, num_of_students)
    table_rows: list[str] = []
    table_list: list[str] = []

    # Create a format for reading the tables and their rows
    for sublist in question_list:

        if (len(sublist) < 1):

            continue

        for index, item in enumerate(sublist):

            if index == 1:

                table_rows.append(["================"])
                
            for sub_item in textwrap.wrap(item):
                
                table_rows.append([sub_item])
        
        table_list.append(table.make_table(table_rows, centered=False))

        table_rows.clear()

    canvas_wb.close()

    # print("\n".join(table_list))

    # Convert into a list of rows to print
    rows_list = "\n".join(table_list).split("\n")
    rows_list.append("===End of List===")
    
    # Add interactivity to scroll up or down using j or k
    curr_row: int = 0
    term = b.Terminal()
    do_clear = True

    while True:

        if do_clear:

            print(term.clear())


            print(term.move(0, 0))

            print_rows(term, rows_list, curr_row)

            print(term.move(term.width - 1, term.height - 1))
            print("\r: ", end="")
        
        do_clear = True

        with term.cbreak():

            code = term.inkey().lower()

            if code == 'q':

                print(term.clear())
                print("Thank you for using Canvas Feedback Reader!")
                break

            elif code == 'j':

                if curr_row < len(rows_list) - term.height + 2:
                    
                    curr_row += 1

                else:

                    do_clear = False

            elif code == 'k':

                if curr_row > 0:

                    curr_row -= 1

                else:

                    do_clear = False

    pass

# The driver call for the program
if __name__ == "__main__":

    main()